#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa contratos a partir de planilha Excel (.xlsx).

Planilha modelo (2 abas):
  - padrao   → valores comuns a todos (template, produto, financeiro, contratada)
  - contratos → uma linha por cliente

Uso via gerar_contrato.py:
  python gerar_contrato.py --gerar-modelo-excel
  python gerar_contrato.py --excel entrada/lote.xlsx
  python gerar_contrato.py --excel entrada/lote.xlsx --dry-run
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise ImportError(
        "Instale openpyxl: pip install openpyxl  (ou pip install -r requirements.txt)"
    ) from exc


COLUNAS_CONTRATOS = [
    ("ativo", "S = gera / N = ignora a linha"),
    ("nome_contrato", "Rótulo interno (opcional)"),
    ("client_id", "Se já existe no Onety; vazio = cria cliente novo"),
    ("tipo_cliente", "empresa | pessoa_fisica"),
    ("razao_social", "Nome / razão social do cliente"),
    ("cpf_cnpj", "CPF ou CNPJ do cliente"),
    ("email", "E-mail do cliente"),
    ("telefone", "Telefone com DDD"),
    ("endereco", "Rua / logradouro"),
    ("numero", "Número"),
    ("complemento", "Complemento (opcional)"),
    ("bairro", "Bairro"),
    ("cidade", "Cidade"),
    ("estado", "UF (ex: SP)"),
    ("cep", "CEP"),
    ("signatario_nome", "Quem assina como CONTRATANTE"),
    ("signatario_cpf", "CPF da pessoa (não use CNPJ aqui)"),
    ("signatario_email", "E-mail do signatário (vazio = usa e-mail do cliente)"),
    ("signatario_telefone", "Telefone do signatário (vazio = usa do cliente)"),
    ("signatario_funcao", "Ex: Assinar como contratante"),
    ("quantidade_licencas", "Ex: 5 → custom.quantidade_de_cnpjs"),
    ("valor_mensal", "Valor recorrente mensal (vazio = usa aba padrao)"),
    ("parcelas", "Meses / parcelas (vazio = usa aba padrao)"),
    ("valor_total", "Total do contrato (vazio = valor_mensal x parcelas)"),
    ("plano_contratado", "Ex: Start (vazio = usa aba padrao)"),
    ("meses_promocionais", "Ex: 0"),
    ("expires_at", "Validade do link YYYY-MM-DD (vazio = calcula)"),
    ("start_at", "Início da vigência YYYY-MM-DD"),
    ("end_at", "Fim da vigência YYYY-MM-DD"),
    ("template_id", "Sobrescreve o da aba padrao (opcional)"),
    ("produto_id", "Sobrescreve o da aba padrao (opcional)"),
    ("produto_nome", "Sobrescreve o da aba padrao (opcional)"),
    ("categoria_id", "Financeiro (opcional)"),
    ("sub_categoria_id", "Financeiro (opcional)"),
    ("conta_api_id", "Conta cobrança (opcional)"),
]

# Padrão FOCO MEI (empresa 785) — LICENÇA FOCOMEI / Start
# Na planilha, em geral só muda a aba "contratos" (cliente + valor se diferente).
CHAVES_PADRAO = [
    ("template_id", "254"),
    ("produto_id", "587"),
    ("produto_nome", "Start"),
    ("produto_tipo", "mensal"),
    ("produto_descricao", "Pacote Start — até 5 CNPJs MEI"),
    ("valor_mensal", "100"),
    ("parcelas", "12"),
    ("plano_contratado", "Start"),
    ("meses_promocionais", "0"),
    ("expires_dias", "2"),
    ("meses_vigencia", "12"),
    ("categoria_id", "12561"),
    ("sub_categoria_id", "86447"),
    ("conta_api_id", "669"),
    ("contratada_nome", "Jhonata Matias Mata Campos"),
    ("contratada_email", "jhonata.matias@contabhub.com.br"),
    ("contratada_cpf", "110.657.647-07"),
    ("contratada_telefone", "22998771692"),
    ("contratada_funcao", "Assinar como contratada"),
    ("signatario_funcao_padrao", "Assinar como contratante"),
]


def _cel(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _num(v: Any, default: float | None = None) -> float | None:
    s = _cel(v).replace(",", ".")
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _int(v: Any, default: int | None = None) -> int | None:
    n = _num(v, None)
    if n is None:
        return default
    return int(n)


def _data_iso(v: Any) -> str:
    s = _cel(v)
    if not s:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # dd/mm/yyyy
    if "/" in s and len(s) >= 8:
        parts = s.split("/")
        if len(parts) == 3:
            d, m, y = parts
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    return s[:10]


def _ler_padrao(ws) -> dict[str, str]:
    """Aba padrao: col A = chave, col B = valor."""
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        chave = _cel(row[0]).lower()
        valor = _cel(row[1]) if len(row) > 1 else ""
        out[chave] = valor
    return out


def _ler_contratos(ws) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_cel(h).lower() for h in rows[0]]
    # linha 2 pode ser descrição — se a 1ª célula for "S" / "N" / número / texto de dado, é dado
    start = 1
    if len(rows) > 1:
        primeira = _cel(rows[1][0]).upper() if rows[1] else ""
        # Se header tem "ativo" e linha 2 parece legenda (texto longo), pula
        if headers and headers[0] == "ativo" and primeira not in ("S", "N", "SIM", "NAO", "NÃO", ""):
            # pode ser exemplo S ou legenda — se tiver muitas palavras, é legenda
            if len(_cel(rows[1][1] if len(rows[1]) > 1 else "")) > 40 or primeira.startswith("S ="):
                start = 2

    itens: list[dict[str, str]] = []
    for row in rows[start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {headers[i]: _cel(row[i]) if i < len(row) else "" for i in range(len(headers))}
        itens.append(item)
    return itens


def _merge(row: dict[str, str], padrao: dict[str, str], chave: str) -> str:
    return (row.get(chave) or padrao.get(chave) or "").strip()


def linha_para_spec(row: dict[str, str], padrao: dict[str, str]) -> dict[str, Any] | None:
    ativo = (row.get("ativo") or "S").strip().upper()
    if ativo in ("N", "NAO", "NÃO", "0", "FALSE", "F"):
        return None

    razao = row.get("razao_social") or row.get("nome") or ""
    if not razao and not row.get("client_id"):
        return None

    template_id = _int(_merge(row, padrao, "template_id"))
    parcelas = _int(_merge(row, padrao, "parcelas"), 12) or 12
    valor_mensal = _num(_merge(row, padrao, "valor_mensal"), 0.0) or 0.0
    valor_total = _num(row.get("valor_total"), None)
    if valor_total is None:
        valor_total = valor_mensal * parcelas

    hoje = date.today()
    start_at = _data_iso(row.get("start_at")) or hoje.isoformat()
    meses_vig = _int(padrao.get("meses_vigencia"), parcelas) or parcelas
    if row.get("end_at"):
        end_at = _data_iso(row.get("end_at"))
    else:
        # aproxima: + meses_vigência
        y, m = hoje.year, hoje.month + meses_vig
        while m > 12:
            y += 1
            m -= 12
        end_at = date(y, m, min(hoje.day, 28)).isoformat()

    expires_at = _data_iso(row.get("expires_at"))
    if not expires_at:
        dias = _int(padrao.get("expires_dias"), 2) or 2
        expires_at = (hoje + timedelta(days=dias)).isoformat()

    tipo = (row.get("tipo_cliente") or "empresa").strip().lower()
    if tipo in ("pj", "pessoa_juridica", "juridica"):
        tipo = "empresa"
    if tipo in ("pf", "fisica"):
        tipo = "pessoa_fisica"

    email = row.get("email") or ""
    telefone = row.get("telefone") or ""
    sig_nome = row.get("signatario_nome") or razao
    sig_email = row.get("signatario_email") or email
    sig_tel = row.get("signatario_telefone") or telefone
    sig_funcao = (
        row.get("signatario_funcao")
        or padrao.get("signatario_funcao_padrao")
        or "Assinar como contratante"
    )

    signatories: list[dict[str, Any]] = [
        {
            "name": sig_nome,
            "email": sig_email,
            "cpf": row.get("signatario_cpf") or "",
            "telefone": sig_tel,
            "funcao_assinatura": sig_funcao,
        }
    ]

    # Contratada (signatário 2) se preenchida na aba padrão
    if padrao.get("contratada_nome") and padrao.get("contratada_email"):
        signatories.append(
            {
                "name": padrao["contratada_nome"],
                "email": padrao["contratada_email"],
                "cpf": padrao.get("contratada_cpf") or "",
                "telefone": padrao.get("contratada_telefone") or "",
                "funcao_assinatura": padrao.get("contratada_funcao")
                or "Assinar como contratada",
            }
        )

    produto_id = _int(_merge(row, padrao, "produto_id"))
    produto_nome = _merge(row, padrao, "produto_nome") or "Produto"
    produto_tipo = padrao.get("produto_tipo") or "mensal"
    qtd_lic = _cel(row.get("quantidade_licencas")) or "1"

    plano = _merge(row, padrao, "plano_contratado")
    meses_promo = _merge(row, padrao, "meses_promocionais") or "0"

    nome_contrato = row.get("nome_contrato") or razao.replace(" ", "_")[:60]

    spec: dict[str, Any] = {
        "nome": nome_contrato,
        "modo": "html",
        "template_id": template_id,
        "expires_at": expires_at,
        "start_at": start_at,
        "end_at": end_at,
        "valor": valor_total,
        "valor_recorrente": valor_mensal,
        "cliente": {
            "tipo": tipo,
            "nome": razao,
            "email": email,
            "cpf_cnpj": row.get("cpf_cnpj") or "",
            "telefone": telefone,
            "endereco": row.get("endereco") or "",
            "numero": row.get("numero") or "",
            "complemento": row.get("complemento") or "",
            "bairro": row.get("bairro") or "",
            "cidade": row.get("cidade") or "",
            "estado": row.get("estado") or row.get("uf") or "",
            "cep": row.get("cep") or "",
        },
        "signatories": signatories,
        "produtos_dados": [
            {
                "id": produto_id,
                "nome": produto_nome,
                "descricao": padrao.get("produto_descricao") or produto_nome,
                "tipo": produto_tipo,
                "quantidade": 1,
                "valor_de_venda": valor_mensal,
                "valor_total": valor_total,
                "parcelas": parcelas,
            }
        ],
        "custom": {
            "plano_contratado": plano,
            "quantidade_de_cnpjs": qtd_lic,
            "meses_promocionais": meses_promo,
        },
    }

    client_id = _int(row.get("client_id"))
    if client_id:
        spec["client_id"] = client_id
    else:
        spec["criar_cliente"] = True

    fin: dict[str, int] = {}
    for k in ("categoria_id", "sub_categoria_id", "conta_api_id"):
        val = _int(_merge(row, padrao, k))
        if val is not None:
            fin[k] = val
    if fin:
        spec["financeiro"] = fin

    return spec


def planilha_para_specs(caminho: Path) -> list[dict[str, Any]]:
    wb = load_workbook(caminho, data_only=True)
    if "padrao" not in wb.sheetnames:
        raise ValueError("A planilha precisa ter a aba 'padrao' (valores comuns).")
    if "contratos" not in wb.sheetnames:
        raise ValueError("A planilha precisa ter a aba 'contratos' (uma linha por cliente).")

    padrao = _ler_padrao(wb["padrao"])
    linhas = _ler_contratos(wb["contratos"])
    specs: list[dict[str, Any]] = []
    for i, row in enumerate(linhas, 1):
        try:
            spec = linha_para_spec(row, padrao)
        except Exception as exc:
            raise ValueError(f"Linha {i} da aba contratos: {exc}") from exc
        if spec:
            specs.append(spec)
    if not specs:
        raise ValueError("Nenhuma linha ativa encontrada na aba contratos.")
    return specs


def gerar_modelo_excel(destino: Path) -> Path:
    """Cria planilha modelo pronta para preencher."""
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── aba padrao ──
    ws_p = wb.active
    ws_p.title = "padrao"
    ws_p.append(["chave", "valor", "descricao"])
    descricoes = {
        "template_id": "ID do modelo (python gerar_contrato.py --listar-modelos)",
        "produto_id": "ID do produto no Onety",
        "produto_nome": "Nome do produto no contrato",
        "produto_tipo": "mensal | anual | avulso",
        "produto_descricao": "Texto descritivo",
        "valor_mensal": "Valor recorrente padrão",
        "parcelas": "Quantidade de meses / parcelas",
        "plano_contratado": "custom.plano_contratado",
        "meses_promocionais": "custom.meses_promocionais",
        "expires_dias": "Dias até o link expirar (se expires_at vazio)",
        "meses_vigencia": "Usado para calcular end_at se vazio",
        "categoria_id": "Financeiro (opcional)",
        "sub_categoria_id": "Financeiro (opcional)",
        "conta_api_id": "Conta de cobrança (opcional)",
        "contratada_nome": "Quem assina como CONTRATADA",
        "contratada_email": "E-mail da contratada",
        "contratada_cpf": "CPF da contratada",
        "contratada_telefone": "Telefone da contratada",
        "contratada_funcao": "Assinar como contratada",
        "signatario_funcao_padrao": "Função padrão do contratante",
    }
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_p[1]:
        cell.fill = header_fill
        cell.font = header_font

    for chave, valor in CHAVES_PADRAO:
        ws_p.append([chave, valor, descricoes.get(chave, "")])

    ws_p.column_dimensions["A"].width = 28
    ws_p.column_dimensions["B"].width = 36
    ws_p.column_dimensions["C"].width = 55

    # ── aba contratos ──
    ws_c = wb.create_sheet("contratos")
    headers = [c[0] for c in COLUNAS_CONTRATOS]
    hints = [c[1] for c in COLUNAS_CONTRATOS]
    ws_c.append(headers)
    ws_c.append(hints)
    for cell in ws_c[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    hint_fill = PatternFill("solid", fgColor="D6EAF8")
    for cell in ws_c[2]:
        cell.fill = hint_fill
        cell.font = Font(italic=True, color="555555", size=9)
        cell.alignment = Alignment(wrap_text=True)

    # Linha de exemplo — só cliente; o resto vem da aba padrao
    exemplo = {
        "ativo": "S",
        "nome_contrato": "Exemplo_Cliente",
        "client_id": "",
        "tipo_cliente": "empresa",
        "razao_social": "Empresa Exemplo LTDA",
        "cpf_cnpj": "00.000.000/0001-00",
        "email": "contato@exemplo.com",
        "telefone": "11999999999",
        "endereco": "Rua Exemplo",
        "numero": "100",
        "complemento": "",
        "bairro": "Centro",
        "cidade": "São Paulo",
        "estado": "SP",
        "cep": "01000-000",
        "signatario_nome": "Nome de Quem Assina",
        "signatario_cpf": "000.000.000-00",
        "signatario_email": "",
        "signatario_telefone": "",
        "signatario_funcao": "",
        "quantidade_licencas": "5",
        "valor_mensal": "",
        "parcelas": "",
        "valor_total": "",
        "plano_contratado": "",
        "meses_promocionais": "",
        "expires_at": "",
        "start_at": "",
        "end_at": "",
        "template_id": "",
        "produto_id": "",
        "produto_nome": "",
        "categoria_id": "",
        "sub_categoria_id": "",
        "conta_api_id": "",
    }
    ws_c.append([exemplo.get(h, "") for h in headers])

    for i, _ in enumerate(headers, 1):
        ws_c.column_dimensions[get_column_letter(i)].width = 18
    ws_c.row_dimensions[1].height = 22
    ws_c.row_dimensions[2].height = 40
    ws_c.freeze_panes = "A3"

    # ── aba ajuda ──
    ws_h = wb.create_sheet("ajuda")
    ajuda = [
        ["Como usar esta planilha"],
        [""],
        ["Aba padrao já vem pronta para FOCO MEI (empresa 785):"],
        ["  modelo 254 LICENÇA FOCOMEI | produto Start 587 | R$ 100 x 12"],
        ["  financeiro 12561 / 86447 / conta 669 | contratada = Jhonata"],
        [""],
        ["1. Em geral NÃO mexa na aba padrao."],
        ["2. Na aba contratos, uma linha = um cliente (ativo=S)."],
        ["3. Preencha só: empresa, CNPJ, endereço, signatário (nome+CPF), e-mail, telefone,"],
        ["   quantidade_licencas. Valor mensal só se for diferente de R$ 100."],
        ["4. client_id vazio → cria o cliente no Onety."],
        ["5. signatario_cpf = CPF de pessoa (NUNCA CNPJ)."],
        ["6. Rode:"],
        ['   python gerar_contrato.py --excel entrada/meu_lote.xlsx --dry-run'],
        ['   python gerar_contrato.py --excel entrada/meu_lote.xlsx'],
        [""],
        ["Dica: copie esta planilha e renomeie (ex.: lote_marco.xlsx)."],
    ]
    for linha in ajuda:
        ws_h.append(linha)
    ws_h["A1"].font = Font(bold=True, size=14)
    ws_h.column_dimensions["A"].width = 90

    wb.save(destino)
    return destino
